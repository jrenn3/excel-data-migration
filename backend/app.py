from flask import Flask, request, send_file, jsonify, make_response
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
import tempfile
import os
import time
import uuid

app = Flask(__name__)
CORS(app)

progress_store = {}

def locate_sheet(workbook, sheet_name):
    for sheet in workbook.worksheets:
        if sheet.title == sheet_name:
            return sheet
    raise ValueError(f"No '{sheet_name}' tab found in {workbook}.")

def copy_data(source_sheet, target_sheet, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            old_cell = source_sheet.cell(row=row, column=col)
            new_cell = target_sheet.cell(row=row, column=col)
            
            # Check for the custom named formula and replace it
            if isinstance(old_cell.value, ArrayFormula) and old_cell.value.text == '=EndDayOfCurrentMonth':
                new_cell.value = '=EndOfCurrentMonth'
            else:
                new_cell.value = old_cell.value
                        
            new_cell.number_format = old_cell.number_format # copy old number format over

def apply_data_validation(sheet, validation_range, source_range):
    data_validation = DataValidation(
        type="list",
        formula1=source_range,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True
    )
    data_validation.add(validation_range)
    sheet.add_data_validation(data_validation)

def migrate_inputs(old_workbook, new_workbook, target_sheet, start_row, end_row, start_col, end_col): #, validation_range=None, source_range=None

    ws_old = locate_sheet(old_workbook, target_sheet)
    ws_new = locate_sheet(new_workbook, target_sheet)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            old_cell = ws_old.cell(row=row, column=col)
            new_cell = ws_new.cell(row=row, column=col)
            
            # Check for the custom named formula and replace it
            if isinstance(old_cell.value, ArrayFormula) and old_cell.value.text == '=EndDayOfCurrentMonth':
                new_cell.value = '=EndOfCurrentMonth'
            else:
                new_cell.value = old_cell.value
                        
            new_cell.number_format = old_cell.number_format # copy old number format over

    # if validation_range and source_range:
    #     apply_data_validation(ws_new, validation_range, source_range)

def migrate_adhoc(old_workbook, new_workbook):
    # Locate the start and end tabs
    sheet_names_old = [sheet.title for sheet in old_workbook.worksheets]
    sheet_names_new = [sheet.title for sheet in new_workbook.worksheets] 
    
    if "AD HOC→" not in sheet_names_old or "FOOTNOTES→" not in sheet_names_old:
        raise ValueError("FOOTNOTES→ or AD HOC→ tabs not found in the workbook.")
    
    start_index_old = sheet_names_old.index("AD HOC→")
    end_index_old = sheet_names_old.index("FOOTNOTES→")

    start_index_new = sheet_names_new.index("AD HOC→")

    # Get all sheets between start_tab and end_tab (exclusive) in the old workbook
    intermediate_sheets = sheet_names_old[start_index_old + 1:end_index_old]

    # Insert sheets between "AD HOC→" and "FOOTNOTES→" in the new workbook
    for i, sheet_name in enumerate(intermediate_sheets):
        ws_old = locate_sheet(old_workbook, sheet_name)
        ws_new = new_workbook.create_sheet(title=sheet_name)

        # Copy all data from the old sheet to the new sheet
        for row in ws_old.iter_rows():
            for cell in row:
                new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
                new_cell.number_format = cell.number_format  # Copy number format

        # Move the new sheet to the correct position in the new workbook
        new_workbook._sheets.remove(ws_new)  # Temporarily remove the sheet
        new_workbook._sheets.insert(start_index_new + 1 + i, ws_new)  # Insert at the correct position

@app.route('/upload', methods=['POST']) # creates endpoint for file upload
def upload():
    upload_id = request.form.get('upload_id') or str(uuid.uuid4())
    print('DEVNOTE upload_id in /upload endpoint:', upload_id)
    progress_store[upload_id] = 0  # 0%

    def update_progress(pct, message):
        progress_store[upload_id] = {'progress': pct, 'message': message}

    #--LOAD WORKBOOK--

    #check if the file was sent correctly
    if 'file' not in request.files: # grabs file from the request via the key 'file'
        update_progress(100, 'Error')
        return 'No file part', 400

    file = request.files['file']
    if file.filename == '':
        update_progress(100, 'Error')
        return 'No selected file', 400

    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp:
        file.save(tmp.name)
        uploaded_path = tmp.name

    update_progress(30, 'Excel bot pouring coffee...')  # File saved

    wb_old = load_workbook(uploaded_path)
    template_path = next((f for f in os.listdir('.') if f.endswith('.xlsm')), None) # loads the template file via extention search
    if not template_path:
        raise FileNotFoundError("No .xlsm template file found in the current directory.")
    
    update_progress(50, 'Copying data like a caffeinated intern...')  # Old file read
    
    wb_new = load_workbook(template_path, keep_vba=True)
        
    #--MIGRATION LOGIC--

    try:
        migrate_inputs(wb_old, wb_new, 'Assets', 3, 99, 2, 5) 
        migrate_inputs(wb_old, wb_new, 'Credit Cards', 3, 24, 2, 6)
        migrate_inputs(wb_old, wb_new, 'Loans', 3, 99, 2, 3)
        migrate_inputs(wb_old, wb_new, 'Loyalty Points & Miles', 3, 53, 2, 6)
        migrate_inputs(wb_old, wb_new, 'Recurring', 3, 53, 2, 7) #todo: map to new category names 
        migrate_inputs(wb_old, wb_new, 'Precedents', 3, 99, 2, 5) #todo: map to new category names 
        migrate_inputs(wb_old, wb_new, 'Changes', 3, 23, 2, 7) #todo: map to new category names 
        migrate_inputs(wb_old, wb_new, 'Planned', 3, 24, 2, 6) #todo: map to new category names 
        migrate_inputs(wb_old, wb_new, 'Blanket', 2, 9, 3, 3) #todo: map to new category names 
        
        apply_data_validation(wb_new['Assets'], "$B$4:$B$99", "'Data Validation'!$A$2:$A$99") #Assets validation
        apply_data_validation(wb_new['Recurring'], "$B$4:$B$99", "'Data Validation'!$B$2:$B$99") #Line items validation
        apply_data_validation(wb_new['Recurring'], "$E$4:$E$99", "'Data Validation'!$C$2:$C$99") #Recurrance base validation
        apply_data_validation(wb_new['Precedents'], "$B$4:$B$99", "'Data Validation'!$B$2:$B$99") #Line items validation
        apply_data_validation(wb_new['Precedents'], "$D$4:$D$99", "'Data Validation'!$B$2:$B$99") #Line items validation for Dependant-on column
        apply_data_validation(wb_new['Changes'],  "$B$4:$B$99", "'Data Validation'!$B$2:$B$99") #Line items validation
        apply_data_validation(wb_new['Changes'],  "$E$4:$E$99", "'Data Validation'!$C$2:$C$99") #Recurrance base validation
        apply_data_validation(wb_new['Planned'],  "$B$4:$B$99", "'Data Validation'!$B$2:$B$99") #Line items validation
        apply_data_validation(wb_new['Planned'],  "$D$4:$D$99", "'Data Validation'!$D$2:$D$99") #Account validation

        update_progress(75, 'Data flowing like coffee on Monday...')  # New file populated

        migrate_adhoc(wb_old, wb_new)

        for sheet in wb_new.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('=') and '{' in cell.value:
                        # Remove curly brackets from the formula
                        cell.value = cell.value.replace('{', '').replace('}', '')

        #--SAVE AND DELIVER FILE--
        output_path = tempfile.mktemp(suffix='.xlsm')
        wb_new.save(output_path)

        update_progress(100, 'Excel bot giving the model a pep talk...')  # Saved file

        response = make_response(send_file(output_path,
                                        as_attachment=True,
                                        mimetype='application/vnd.ms-excel.sheet.macroEnabled.12'))
        response.headers['X-Upload-Id'] = upload_id
        return response
    
    except Exception as e:
        update_progress(100, "Error")
        response = jsonify({'error': str(e)})
        response.headers['X-Upload-Id'] = upload_id
        return response, 500
    
    finally:
        os.unlink(uploaded_path)
        # os.unlink(output_path)

@app.route('/progress/<upload_id>', methods=['GET'])
def progress(upload_id):
    info  = progress_store.get(upload_id, {'progress': 0, 'message': 'Waiting to start...'})
    return jsonify(info)

# --START SERVER--
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))  # fallback to 5000 for local dev
    app.run(host='0.0.0.0', port=port)