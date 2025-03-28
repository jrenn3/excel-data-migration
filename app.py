from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook
import tempfile
import os
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

@app.route('/upload', methods=['POST']) # creates endpoint for file upload
def upload():

    print('DEVNOTE endpoint hit')

    #check if the file was sent correctly
    if 'file' not in request.files: # grabs file from the request via the key 'file'
        return 'No file part', 400

    file = request.files['file']
    if file.filename == '':
        return 'No selected file', 400

    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp:
        file.save(tmp.name)
        uploaded_path = tmp.name

    print('DEVNOTE temp file saved')

    # TRANSFORMATION LOGIC - currently a mockup
    try:

        print('DEVNOTE transformation logic started')

        # Load uploaded file and extract user inputs
        wb_old = load_workbook(uploaded_path)
        ws_old = wb_old.active
        user_name = ws_old['A1'].value # if 'A1' in ws_old else 'Anonymous'

        # Load blank template to populate
        template_path = 'public/TEST_DRAFT_TEMPLATE_FUNdsForecast_ByMoe_v6_2025_03_27.xlsm'  # pre-saved new version
        wb_new = load_workbook(template_path, keep_vba=True)
        ws_new = wb_new.active

        print('DEVNOTE template loaded')

        # Insert user data (mock)
        ws_new['A1'] = user_name  # Assuming B2 is where user's name goes in the new version

        # Save output file
        output_path = tempfile.mktemp(suffix='.xlsm')
        wb_new.save(output_path)

        print('DEVNOTE output file saved')

        return send_file(output_path,
                         as_attachment=True,
                         download_name='updated_template.xlsm',
                         mimetype='application/vnd.ms-excel.sheet.macroEnabled.12')
    except Exception as e:
        return f'Error processing file: {str(e)}', 500
    finally:
        os.unlink(uploaded_path)

if __name__ == '__main__':
    app.run(debug=True)
